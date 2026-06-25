from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = {
    "fr": ["N° d'ordre", "Date", "Type", "Référence", "Objet", "Expéditeur", "Projet", "Destinataire", "Date remise", "Statut"],
    "en": ["Order No.", "Date", "Type", "Reference", "Subject", "Sender", "Project", "Recipient", "Delivery date", "Status"],
}
FIELDS = ["no_ordre", "date_enregistrement", "type_document", "reference", "objet", "expediteur", "projet", "destinataire", "date_remise_destinataire", "dernier_statut"]
WIDTHS = [15, 12, 16, 16, 42, 22, 22, 22, 14, 12]


def _style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="075095")
        cell.alignment = Alignment(vertical="center")


def build_journal_xlsx(records: list[dict], lang: str = "fr") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal"
    ws.append(HEADERS.get(lang, HEADERS["fr"]))
    _style_header(ws)
    for r in records:
        ws.append([r.get(f) for f in FIELDS])
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


AUDIT_HEADERS = ["Date / heure", "Utilisateur", "Action", "Entité", "Cible", "IP"]
AUDIT_WIDTHS = [20, 26, 22, 16, 40, 16]


def build_audit_xlsx(items: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal d'activité"
    ws.append(AUDIT_HEADERS)
    _style_header(ws)
    for it in items:
        at = it.get("at")
        ws.append([
            at.strftime("%d/%m/%Y %H:%M:%S") if at else "",
            it.get("actor_full_name") or it.get("actor_username") or "—",
            it.get("action"),
            it.get("entity") or "",
            it.get("entity_id") or "",
            it.get("ip") or "",
        ])
    for i, w in enumerate(AUDIT_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
