from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.deps import get_db
from app.core.security import hash_password
from app.main import app
from app.models.user import AppUser


@pytest.fixture
def client(Session):
    def _override():
        with Session() as s:
            yield s

    app.dependency_overrides[get_db] = _override
    with Session() as s:
        s.add(AppUser(username="amel", full_name="Amel",
                      password_hash=hash_password("pw"), role="clerk"))
        s.commit()
    yield TestClient(app)
    app.dependency_overrides.clear()


def _auth(client):
    tok = client.post("/auth/login", json={"username": "amel", "password": "pw"}).json()["access_token"]
    return {"Authorization": f"Bearer {tok}"}


def _seed(client, h, n=3):
    for i in range(n):
        client.post("/registers/entree/documents",
                    json={"type_document": "Lettre", "objet": f"Objet {i}", "expediteur": "ONAS"},
                    headers=h)


def test_export_xlsx(client):
    h = _auth(client)
    _seed(client, h, 3)

    r = client.get("/export/journal.xlsx?register=entree", headers=h)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "N° d'ordre"   # FR header
    assert ws.max_row == 4                                  # header + 3 rows
    assert str(ws.cell(row=2, column=1).value).startswith("BOE")


def test_report_data_filtered(client):
    h = _auth(client)
    _seed(client, h, 3)

    r = client.get("/reports/journal-data?register=entree&q=Objet%201", headers=h)
    assert r.status_code == 200
    body = r.json()
    assert body["register"] == "E"
    assert body["count"] == 1
    assert body["items"][0]["objet"] == "Objet 1"
